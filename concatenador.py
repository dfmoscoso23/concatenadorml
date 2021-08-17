#Backstage
import openpyxl as op
import pandas as pd
import obtenerTokenmodulo as otm
from tkinter import*
from tkinter import ttk
from tkinter import messagebox as mb
import requests
import json
import correctortitulo as ct
tok= False 
#Diccionarios
dic_titulo = {}
dic_autor = {}
dic_editorial = {}
dic_precio = {}
dic_titulopub = {}
portada = {}
contraportada = {}
paginaextra1 = {}
paginaextra2 = {}
paginaextra3 = {}
paginaextra4 = {}
dic_tema = {}
apellido = {}
imagenes = {}
lista_publicados=[]
lista_no_publicados=[]
dic_no_publicados={}
isbns = []
isbn_depurado = []
en_la_base = []
no_en_la_base = []
corregidos={}
dic_real_isbn={}
corregidos_isbn_concat={}
corregidos_precio_concat={}
global excel
excel=False
global csvcon
csvcon=False
a=0

def revisar_tokenpresxistente():
	global tok
	tok=otm.revisar_token_activo()
	if tok==False:
		otm.obtenertoken()
	else:	
		pass
def funerariat(tit, isbnd):
	"""
	colon=""
	if ',' in tit:
		titu = tit.split(',')
		arti = titu [1].strip(" ")
		titul = titu[0].strip(" ")
		largo_arit=len(arti)
		if largo_arit > 3:
				if arti[2]==" ":
					artic= arti[0]+arti[1]
				elif arti[2]=="s" or arti[2]=="o" or arti[2]=="a":
					artic= arti[0]+arti[1]+arti[2]
					colon=arti[3:]
				elif arti[2]=="/":
					artic= arti[0]+arti[1]
				else:
					artic= arti[0]+arti[1]
					colon=arti[3:]			
		else:
			if arti[1]==" ":
				artic=	arti[0]
			elif arti[1]=="a" or arti[1]=="l" or arti[1]=="o":
				artic= arti[0]+arti[1]
			else:
				artic=""				
		if len(colon)>0:
			titulo = artic + " " + titul+" "+colon
		else:
			titulo = artic + " " + titul
	else:
		titulo = tit
	tituloo = titulo.title()
	titulod = tituloo.strip()
	dic_titulo[isbnd] = titulod
	"""
	dic_titulo[isbnd]=ct.corrector_titulo(tit)
def funerariaa(aut, isbnd):
	"""
	if ',' in aut:
		auto = aut.split(',')
		artic = auto [1].strip(" ")
		ape = auto[0].strip(" ")
		autor = artic + " " + ape
		apellido[isbnd] = ape.title()
	else:
		autor = aut.strip("	")
		apellido[isbnd] = autor.title()
	autort = autor.strip(" ")
	autord = autort.title()
	#ct.corrector_autor(aut)
	dic_autor[isbnd] = autord
	"""
	autor,apell=ct.corrector_autor(aut)
	apellido[isbnd] = apell
	dic_autor[isbnd] = autor

def funerariae(edit, isbnd):
	editora = edit.title()
	editorial = editora.strip(" ")
	dic_editorial[isbnd] = editorial
def funerariap(pre, isbnd):
	dic_precio[isbnd] =str(pre)	

def concatenadopub(isbnd):
	dic_titulopub[isbnd] = dic_titulo[isbnd] + " - "+ apellido[isbnd] + " - " + dic_editorial[isbnd]

def funerariatema(isbnd, tem):
	dic_tema[isbnd] = tem.title()
def desplegar_copiableaexcel():
	toplevel_arbol=Toplevel(window)
	toplevel_arbol.title("Publicaciones")
	cuadro_resultado_frame=Frame(toplevel_arbol)
	cuadro_resultado_frame.pack()
	cuadro_resultado = Text(cuadro_resultado_frame, width=75, height=25)
	cuadro_resultado.grid(column=1, row=3, padx=5, pady=5)

	scroll = Scrollbar(cuadro_resultado_frame, command=cuadro_resultado.yview)
	scroll.grid(column=2, row=3, sticky="nsew")
	cuadro_resultado.config(yscrollcommand=scroll.set)
	for isbnd in en_la_base:
		cuadro_resultado.insert(
			END, dic_titulopub[isbnd] + "," +
			isbnd + "," + imagenes[isbnd] + "," +
			isbnd + "," + "1" + "," + dic_precio[isbnd] + "," + "Nuevo" + "," + "des" + "," + " " + "," +
			"Clásica" + "," + "Mercado Envíos | Mercado Envíos Flex" + "," + "A cargo del comprador" + "," +
			"Acepto" + "," + "Garantía del vendedor"  + "," + "1" + "," + "meses"  + "," + "Papel" + "," + 
			dic_tema[isbnd]  + "," + dic_titulo[isbnd] + "," + dic_autor[isbnd] + "," + "Español" + "," + 
			dic_editorial[isbnd] + "," + dic_tema[isbnd] + "\n"
			)

def arbolgeneral():
	a=0
	for isb in en_la_base:
		arb.insert(parent='',index=a, iid=a, text='', values=(dic_titulopub[isb][:59],dic_real_isbn[isb],imagenes[isb],isb,"1",dic_precio[isb],dic_tema[isb],dic_titulo[isb],dic_autor[isb],dic_editorial[isb]))
		a+=1
	arb.pack()	
def busqueda(xrow, isbnd, x):
	#cuadro_resultado.insert(END, isbnd)
	#cuadro_resultado.insert(END, xrow)
	if isbnd==str(xrow):
		en_la_base.append(isbnd)
		global sheet
		global excel_titulo_conc
		global excel_autor_conc
		global excel_editorial_conc
		global excel_precio_conc
		global excel_tema_conc	
		#print(sheet.cell(row=x, column=2).value)
		tit = sheet.cell(row=x, column=excel_titulo_conc).value
		aut = sheet.cell(row=x, column=excel_autor_conc).value
		edit = sheet.cell(row=x, column=excel_editorial_conc).value
		pre = sheet.cell(row=x, column=excel_precio_conc).value
		tem = sheet.cell(row=x, column=excel_tema_conc).value
		funerariat(tit, isbnd)
		funerariaa(aut, isbnd)
		funerariae(edit, isbnd)
		funerariap(pre, isbnd)
		funerariatema(isbnd, tem)
		concatenadopub(isbnd)
		dic_real_isbn[isbnd]=isbnd
def busqudacsv(isbnd):

	#if len(isbnd)!=13:
		#bus_isbn = str(isbnd)+"   "
	#else:
		#bus_isbn= str(isbnd)
	bus_isbn=int(isbnd)		
	global csv
	fila_isbn=csv[csv.isbn==bus_isbn]
	print(fila_isbn)
	global num_tit
	global num_aut
	global num_edit
	global num_pre
	global num_tem
	try:
		tit = fila_isbn.iat[0,int(num_tit)]
		aut = fila_isbn.iat[0,int(num_aut)]
		edit = fila_isbn.iat[0,int(num_edit)]
		pre = fila_isbn.iat[0,int(num_pre)]
		tem = fila_isbn.iat[0,int(num_tem)]	
		funerariat(str(tit), isbnd)
		funerariaa(str(aut), isbnd)
		funerariae(str(edit), isbnd)
		funerariap(pre, isbnd)
		funerariatema(isbnd, str(tem))
		concatenadopub(isbnd)
		dic_real_isbn[isbnd]=isbnd
		en_la_base.append(isbnd)

	except IndexError:
		no_en_la_base.append(isbnd)			
def concatenado(isbnd):
	"""concatena las imágenes de portada y contraportada"""
	if isbnd in portada:
		if isbnd in contraportada:
			if isbnd in paginaextra1:
				if isbnd in paginaextra2:
					if isbnd in paginaextra3:
						if isbnd in paginaextra4:
							imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; " + paginaextra2[isbnd] + "; " + paginaextra3[isbnd] + "; " + paginaextra4[isbnd]# + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
						else:
							imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; " + paginaextra2[isbnd] + "; " + paginaextra3[isbnd]# + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
					else:
						imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; " + paginaextra2[isbnd]# + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
				else:
					imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd]# + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
			else:
				imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd]# + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
				#cuadro_resultado.insert(END, imagenes)
		else:
			imagenes[isbnd] = portada[isbnd]# + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
	#else:
		#imagenes[isbnd] =  "https://i.postimg.cc/B6SMfSSh/001.jpg"

def deconstruirisbns(presibn, listas):
	if presibn[-7] != "0":
		isbn_con_error(presibn)
		#cuadro_resultado.insert(END, "\n" + presibn + " será excluido de la lista final porque no tiene el formato adecuado.\n Recordá que después del ISBN debe incluir 001.jpg\n así se determina la posición de la foto.")
	else:
		if presibn[-21] == "/":
			largo = len(presibn)
			desde = int(largo)-20
			preisbna = presibn[desde:largo]
			isbn = preisbna[0:13]
			resto = preisbna[13:20]
			isbns.append(isbn)
			if resto == "001.jpg":
				portada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "002.jpg":
				contraportada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "003.jpg":
				paginaextra1[isbn] = listas[0]
				listas.pop(0)
			elif resto == "004.jpg":
				paginaextra2[isbn] = listas[0]
				listas.pop(0)
			elif resto == "005.jpg":
				paginaextra3[isbn] = listas[0]
				listas.pop(0)
			elif resto == "006.jpg":
				paginaextra4[isbn] = listas[0]
				listas.pop(0)
		else:
			largo = len(presibn)
			desde = int(largo)-17
			preisbna = presibn[desde:largo]
			isbn = preisbna[0:10]
			resto = preisbna[10:17]
			isbns.append(isbn)
			if resto == "001.jpg":
				portada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "002.jpg":
				contraportada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "003.jpg":
				paginaextra1[isbn] = listas[0]
				listas.pop(0)
			elif resto == "004.jpg":
				paginaextra2[isbn] = listas[0]
				listas.pop(0)
			elif resto == "005.jpg":
				paginaextra3[isbn] = listas[0]
				listas.pop(0)
			elif resto == "006.jpg":
				paginaextra4[isbn] = listas[0]
				listas.pop(0)
def fin():
	mb.showinfo('Aviso', 'Proceso concluido')
def noenlabase():
	for n in no_en_la_base:
		mb.showinfo('Aviso', n + '\n no se encontraban en la base de datos')	

def cargarexcel():
	global excel_concatenador_libro
	libro_ex_conca=excel_concatenador_libro.get()
	global excel_concatenador_hoja
	hoja_ex_conca=excel_concatenador_hoja.get()
	wb = op.load_workbook(libro_ex_conca+'.xlsx')
	global sheet
	sheet = wb[hoja_ex_conca]
	global excel_concatenador_columna_isbn_ent
	global excel_isbn_conc
	excel_isbn_conc = int(excel_concatenador_columna_isbn_ent.get())
	global excel_concatenador_columna_titulo_ent
	global excel_titulo_conc
	excel_titulo_conc = int(excel_concatenador_columna_titulo_ent.get())
	global excel_concatenador_columna_autor_ent
	global excel_autor_conc
	excel_autor_conc = int(excel_concatenador_columna_autor_ent.get())
	global excel_concatenador_columna_editorial_ent
	global excel_editorial_conc
	excel_editorial_conc = int(excel_concatenador_columna_editorial_ent.get())
	global excel_concatenador_columna_precio_ent
	global excel_precio_conc
	excel_precio_conc = int(excel_concatenador_columna_precio_ent.get())
	global excel_concatenador_columna_tema_ent
	global excel_tema_conc
	excel_tema_conc = int(excel_concatenador_columna_tema_ent.get())
	global excel
	excel = True
	toplevel_desdeExcel_concatenador.destroy()
def desdeExcel_concatenador():
	#Toplevel de ingreso de datos
	global toplevel_desdeExcel_concatenador
	toplevel_desdeExcel_concatenador = Toplevel(window)
	toplevel_desdeExcel_concatenador.title("Ingresar Catálogo desde Excel")
	frame_DE = Frame(toplevel_desdeExcel_concatenador)
	frame_DE.pack()
	frame_DE2=Frame(toplevel_desdeExcel_concatenador)
	frame_DE2.pack()
	#Ingresar lista de publicaciones
	excel_concatenador_libro_lab = Label(frame_DE, text="Inserte Libro de publicaciones:")
	excel_concatenador_libro_lab.grid(column=1, row=2)
	excel_concatenador_libro_lab2 = Label(frame_DE, text=".xlsx")
	excel_concatenador_libro_lab2.grid(column=3, row=2)
	global excel_concatenador_libro
	excel_concatenador_libro = Entry(frame_DE, width=15)
	excel_concatenador_libro.grid(column=2, row=2, padx=5, pady=5)
	excel_concatenador_libro.insert(END,"EML")
	excel_concatenador_hoja_lab = Label(frame_DE, text="Inserte nombre de la hoja:")
	excel_concatenador_hoja_lab.grid(column=1, row=3)
	global excel_concatenador_hoja
	excel_concatenador_hoja = Entry(frame_DE, width=15)
	excel_concatenador_hoja.grid(column=2, row=3, padx=5, pady=5)
	excel_concatenador_hoja.insert(END,"EML")
	excel_concatenador_columna_titulo_lab= Label(frame_DE, text="Inserte el número de columna de Título")
	excel_concatenador_columna_titulo_lab.grid(column=1, row=4)
	global excel_concatenador_columna_titulo_ent
	excel_concatenador_columna_titulo_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_titulo_ent.grid(column=2, row=4, padx=5, pady=5)
	excel_concatenador_columna_titulo_ent.insert(END,"2")
	excel_concatenador_columna_autor_lab= Label(frame_DE, text="Inserte el número de columna de Autor:")
	excel_concatenador_columna_autor_lab.grid(column=1, row=5)
	global excel_concatenador_columna_autor_ent
	excel_concatenador_columna_autor_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_autor_ent.grid(column=2, row=5, padx=5, pady=5)
	excel_concatenador_columna_autor_ent.insert(END,"3")
	excel_concatenador_columna_editorial_lab= Label(frame_DE, text="Inserte el número de columna de Editorial:")
	excel_concatenador_columna_editorial_lab.grid(column=1, row=6)
	global excel_concatenador_columna_editorial_ent
	excel_concatenador_columna_editorial_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_editorial_ent.grid(column=2, row=6, padx=5, pady=5)
	excel_concatenador_columna_editorial_ent.insert(END,"4")
	excel_concatenador_columna_isbn_lab= Label(frame_DE, text="Inserte el número de columna de ISBN:")
	excel_concatenador_columna_isbn_lab.grid(column=1, row=7)
	global excel_concatenador_columna_isbn_ent
	excel_concatenador_columna_isbn_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_isbn_ent.grid(column=2, row=7, padx=5, pady=5)
	excel_concatenador_columna_isbn_ent.insert(END,"9")
	excel_concatenador_columna_precio_lab= Label(frame_DE, text="Inserte el número de columna de Precio:")
	excel_concatenador_columna_precio_lab.grid(column=1, row=8)
	global excel_concatenador_columna_precio_ent
	excel_concatenador_columna_precio_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_precio_ent.grid(column=2, row=8, padx=5, pady=5)
	excel_concatenador_columna_precio_ent.insert(END,"29")
	excel_concatenador_columna_tema_lab= Label(frame_DE, text="Inserte el número de columna de Tema:")
	excel_concatenador_columna_tema_lab.grid(column=1, row=9)
	global excel_concatenador_columna_tema_ent
	excel_concatenador_columna_tema_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_tema_ent.grid(column=2, row=9, padx=5, pady=5)
	excel_concatenador_columna_tema_ent.insert(END,"6")
	bot_guardar_catalogo = Button(frame_DE2, text="Guardar", command=cargarexcel)
	bot_guardar_catalogo.pack()
def cargarcsv():
	global Csv_concatenador_libro
	hoja_csv=Csv_concatenador_libro.get()
	global csv
	csv=pd.read_csv(hoja_csv+'.txt',sep=';')
	global num_tit
	num_tit = Csv_concatenador_columna_titulo_ent.get()
	global num_aut
	num_aut = Csv_concatenador_columna_autor_ent.get()
	global num_edit
	num_edit = Csv_concatenador_columna_editorial_ent.get()
	global num_pre
	num_pre = Csv_concatenador_columna_precio_ent.get()
	global num_tem
	num_tem = Csv_concatenador_columna_tema_ent.get()
	global csvcon
	csvcon = True
	toplevel_desdeCsv_concatenador.destroy()
def desdecsv_concatenador():
	#Toplevel de ingreso de datos
	global toplevel_desdeCsv_concatenador
	toplevel_desdeCsv_concatenador = Toplevel(window)
	toplevel_desdeCsv_concatenador.title("Ingresar Catálogo desde Csv")
	frame_DE = Frame(toplevel_desdeCsv_concatenador)
	frame_DE.pack()
	frame_DE2=Frame(toplevel_desdeCsv_concatenador)
	frame_DE2.pack()
	#Ingresar lista de publicaciones
	Csv_concatenador_libro_lab = Label(frame_DE, text="Inserte Libro de publicaciones:")
	Csv_concatenador_libro_lab.grid(column=1, row=2)
	Csv_concatenador_libro_lab2 = Label(frame_DE, text=".csv")
	Csv_concatenador_libro_lab2.grid(column=3, row=2)
	global Csv_concatenador_libro
	Csv_concatenador_libro = Entry(frame_DE, width=15)
	Csv_concatenador_libro.grid(column=2, row=2, padx=5, pady=5)
	Csv_concatenador_libro.insert(END,"EML")
	Csv_concatenador_columna_titulo_lab= Label(frame_DE, text="Inserte el número de columna de Título")
	Csv_concatenador_columna_titulo_lab.grid(column=1, row=4)
	global Csv_concatenador_columna_titulo_ent
	Csv_concatenador_columna_titulo_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_titulo_ent.grid(column=2, row=4, padx=5, pady=5)
	Csv_concatenador_columna_titulo_ent.insert(END,"1")
	Csv_concatenador_columna_autor_lab= Label(frame_DE, text="Inserte el número de columna de Autor:")
	Csv_concatenador_columna_autor_lab.grid(column=1, row=5)
	global Csv_concatenador_columna_autor_ent
	Csv_concatenador_columna_autor_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_autor_ent.grid(column=2, row=5, padx=5, pady=5)
	Csv_concatenador_columna_autor_ent.insert(END,"2")
	Csv_concatenador_columna_editorial_lab= Label(frame_DE, text="Inserte el número de columna de Editorial:")
	Csv_concatenador_columna_editorial_lab.grid(column=1, row=6)
	global Csv_concatenador_columna_editorial_ent
	Csv_concatenador_columna_editorial_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_editorial_ent.grid(column=2, row=6, padx=5, pady=5)
	Csv_concatenador_columna_editorial_ent.insert(END,"3")
	Csv_concatenador_columna_isbn_lab= Label(frame_DE, text="Inserte el número de columna de ISBN:")
	Csv_concatenador_columna_isbn_lab.grid(column=1, row=7)
	global Csv_concatenador_columna_isbn_ent
	Csv_concatenador_columna_isbn_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_isbn_ent.grid(column=2, row=7, padx=5, pady=5)
	Csv_concatenador_columna_isbn_ent.insert(END,"8")
	Csv_concatenador_columna_precio_lab= Label(frame_DE, text="Inserte el número de columna de Precio:")
	Csv_concatenador_columna_precio_lab.grid(column=1, row=8)
	global Csv_concatenador_columna_precio_ent
	Csv_concatenador_columna_precio_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_precio_ent.grid(column=2, row=8, padx=5, pady=5)
	Csv_concatenador_columna_precio_ent.insert(END,"5")
	Csv_concatenador_columna_tema_lab= Label(frame_DE, text="Inserte el número de columna de Tema:")
	Csv_concatenador_columna_tema_lab.grid(column=1, row=9)
	global Csv_concatenador_columna_tema_ent
	Csv_concatenador_columna_tema_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_tema_ent.grid(column=2, row=9, padx=5, pady=5)
	Csv_concatenador_columna_tema_ent.insert(END,"9")
	bot_guardar_catalogo = Button(frame_DE2, text="Guardar", command=cargarcsv)
	bot_guardar_catalogo.pack()

def concatenar():
	global excel
	global csvcon
	if excel == False and csvcon == False:
		mb.showerror("Error","No se ha cargado el catálogo")
	else:		
		lista = ingreso.get()
		listas = lista.split()
		preisbns = listas[:]
		while len(listas)>0:
			for d in preisbns:
				deconstruirisbns(d, listas)
			for i in isbns:
				if i not in isbn_depurado:
					isbn_depurado.append(i)
			for isbnd in isbn_depurado:
				concatenado(isbnd)				
				if excel==True:
					global sheet
					ga = (len(sheet['A'])+1)
					for x in range (1,ga):
						global excel_isbn_conc
						xrow = sheet.cell(row=x, column=excel_isbn_conc).value
						busqueda(xrow, isbnd, x)
					if isbnd not in en_la_base:
						no_en_la_base.append(isbnd)			
				elif csvcon==True:
					busqudacsv(isbnd) 		
				#cuadro_resultado.insert(END, dic_titulopub[isbnd] + "," + isbnd + "," + imagenes[isbnd] + "," + isbnd + "," + dic_titulo[isbnd] + "," + dic_autor[isbnd] + "," + dic_editorial[isbnd])
		arbolgeneral()
		print(no_en_la_base)
		noenlabase()
		fin()
def isbn_con_error(isbn):
	mb.showinfo('Aviso', isbn + ' no pudo ser concatenado por no tener el formato adecuado')



def formato_correcto():
	mb.showinfo(
		'Formato correcto',
		'Los URLs deben ser ingresados en una línea separados por espacios.'
		+ '\n con la forma: http://***/ISBN001.jpg \n' +
		'Donde 001.jpg será la portada, 002.jpg será la contraportada \n'
		+ 'se puede incluir hasta 006.jpg.\n' +
		'El concatenador soporta ISBN 10 y EAN13'
		)
def desplegar_corrector_concat():
	global desplegable_corector
	desplegable_corector = Toplevel(window)
	desplegable_corector.title("Corrector")
	botonera_desplegable = Frame(desplegable_corector)
	botonera_desplegable.pack()
	boton_corregir= Button(botonera_desplegable,text="Guardar",command=corregir)
	boton_corregir.pack()
	frame2=Frame(desplegable_corector)
	frame2.pack(fill=BOTH, expand=1)

	canvas = Canvas(frame2)
	canvas.pack(side=LEFT, fill=BOTH, expand=1)

	scroll = ttk.Scrollbar(frame2, orient=VERTICAL, command=canvas.yview)
	scroll.pack(side=RIGHT, fill=Y)

	canvas.configure(yscrollcommand=scroll.set)
	canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

	panel_corrector = Frame(canvas)
	canvas.create_window((0,0), window=panel_corrector, anchor="nw")
	#panel_corrector= Frame(desplegable_corector)
	#panel_corrector.pack()
	x=0
	for isbn in en_la_base:
		label_titulo_publicacion = Label(panel_corrector, text=dic_titulopub[isbn][:59])
		label_titulo_publicacion.grid(column=0,row=x)
		entry_titulo_publicacion = Entry(panel_corrector, width=50)
		entry_titulo_publicacion.grid(column=0,row=x+1, padx=5, pady=5)
		entry_titulo_publicacion.insert(END, dic_titulopub[isbn])
		corregidos[isbn]=entry_titulo_publicacion
		label_isbn_correccion = Label(panel_corrector, text=dic_real_isbn[isbn])
		label_isbn_correccion.grid(column=1,row=x)
		entry_isbn_correccion = Entry(panel_corrector, width=20)
		entry_isbn_correccion.grid(column=1,row=x+1, padx=5, pady=5)
		entry_isbn_correccion.insert(END, dic_real_isbn[isbn])
		corregidos_isbn_concat[isbn]=entry_isbn_correccion
		label_precio_correccion = Label(panel_corrector, text=dic_precio[isbn])
		label_precio_correccion.grid(column=2,row=x)
		entry_precio_correccion = Entry(panel_corrector, width=10)
		entry_precio_correccion.grid(column=2,row=x+1, padx=5, pady=5)
		entry_precio_correccion.insert(END, dic_precio[isbn])
		corregidos_precio_concat[isbn]=entry_precio_correccion
		x+=2
def corregir():
	for isbn, corregido in corregidos.items():
		item=corregido.get()
		dic_titulopub[isbn]=item
	for isbn, correg in corregidos_isbn_concat.items():
		items=correg.get()
		dic_real_isbn[isbn]=items
	for isbn, corregi in corregidos_precio_concat.items():
		ite=corregi.get()
		dic_precio[isbn]=ite	  	
	for item in arb.get_children():
   		arb.delete(item)	
	arbolgeneral()
	global desplegable_corector
	desplegable_corector.destroy()
def publicar():
	if tok == False:
		mb.showerror("Error", "No hay token Cargado")
	else:	
		for isbnd in en_la_base:
			imag_sub=imagenes[isbnd].split("; ")
			lista_para_subir = []
			for i in imag_sub:
				lista_para_subir.append({"source":i})
				print(i)
			print(lista_para_subir)	
			url="https://api.mercadolibre.com/items"
			token=tok
			headers = {"Authorization": str("Bearer "+token)}
			data= {
			  "title": str(dic_titulopub[isbnd][:59]),
			  "category_id":"MLA412445",
			  "price":dic_precio[isbnd],
			  "currency_id":"ARS",
			  "available_quantity":1,
			  "buying_mode":"buy_it_now",
			  "condition":"used",
			  "listing_type_id":"gold_special",
			  "sale_terms":[
			     {
			        "id":"WARRANTY_TYPE",
			        "value_name":"Garantía del vendedor"
			     },
			     {
			        "id":"WARRANTY_TIME",
			        "value_name":"30 días"
			     }
			  ],
			  "pictures":lista_para_subir,
			  "shipping":{
			  	"mode":"me2",
			  	"local_pick_up": True,
			  	"logistic_type": "xd_drop_off"
			  },
			  "attributes":[
			     {
			        "id":"AUTHOR",
			        "value_name": dic_autor[isbnd]
			     },
			     {
			        "id":"BOOK_GENRE",
			        "value_name": dic_tema[isbnd]
			     },
			     {
			        "id":"BOOK_TITLE",
			        "value_name":dic_titulo[isbnd]
			     },
			     {
			        "id":"FORMAT",
			        "value_name":"Papel"
			     },
			     {
			        "id":"GTIN",
			        "value_name":dic_real_isbn[isbnd]
			     },
			     {
			        "id":"ITEM_CONDITION",
			        "value_name":"Usado"
			     },
			     {
			        "id":"LANGUAGE",
			        "value_name":"Español"
			     },
			     {
			        "id":"NARRATION_TYPE",
			        "value_name":dic_tema[isbnd]
			     },
			     {
			        "id":"PUBLISHER",
			        "value_name":dic_editorial[isbnd]
			     },
			     {
			        "id":"SELLER_SKU",
			        "value_name":isbnd
			     }
			  ]
			}
			req = requests.post(url, headers=headers, json=data)
			if req.status_code == 201:
				lista_publicados.append(isbnd)
			else:
				lista_no_publicados.append(isbnd)
				dic_no_publicados[isbnd] = req.content
		mb.showinfo("Proceso concluido", "Se publicaron "+str(len(lista_publicados))+" items")	
#GUI



def excluidos():
	nv = Toplevel(window)
	cuadro_excluidos = Text(nv, width=75, height=25)
	cuadro_excluidos.pack()
	for isbndn in no_en_la_base:
		cuadro_excluidos.insert(END, isbndn + "," + imagenes[isbndn] + "\n")
def nopubicados():
	nv = Toplevel(window)
	cuadro_excluidos = Text(nv, width=75, height=25)
	cuadro_excluidos.pack()
	for isbndn in lista_no_publicados:
		cuadro_excluidos.insert(END, isbndn + "," + imagenes[isbndn] +"\n")
		print(dic_no_publicados[isbndn])
		print(type(dic_no_publicados[isbndn]))
	for isbndn in lista_no_publicados:			
		bit = dic_no_publicados[isbndn].decode()
		cuadro_excluidos.insert(END, isbndn + "," + bit)

window = Tk()
window.title("Librería Losada")


frame = Frame(window)
frame.grid(column=0, row=0)
botonera_concatenador = Frame(window)
botonera_concatenador.grid(column=1, row=3)

bienvenida = Label(frame, text="Concatenador de imágenes")
bienvenida.grid(column=1, row=0)


ingrese = Label(
	frame,
	text="Ingrese URL de imágenes:"
	)

ingrese.grid(column=0, row=1)

ingreso = Entry(frame, width=75)
ingreso.grid(column=1, row=1, padx=5, pady=5)

boton = Button(botonera_concatenador, text="Concatenar URL", width=15, height=5, command=concatenar)
boton.grid(column=0, row=3,padx=5, pady=5)

boton_excel_concatenador = Button(botonera_concatenador, text="Cargar Excel", width=10, height=1, command=desdeExcel_concatenador)
boton_excel_concatenador.grid(column=0, row=1,padx=5, pady=5)
boton_csv_concatenador = Button(botonera_concatenador, text="Cargar CSV", width=10, height=1, command=desdecsv_concatenador)
boton_csv_concatenador.grid(column=1, row=1,padx=5, pady=5)
boton_toplevel_arbol = Button(botonera_concatenador, text="desplegar copiable", command=desplegar_copiableaexcel)
boton_toplevel_arbol.grid(column=0, row=4,padx=5, pady=5)
boton_de_formato = Button(botonera_concatenador, text="Ver formato correcto", command=formato_correcto)
boton_de_formato.grid(column=0, row=0,padx=5, pady=5)
boton_de_excluidos = Button(botonera_concatenador, text="Ver URLs excluidos", command=excluidos)
boton_de_excluidos.grid(column=1, row=4,padx=5, pady=5)
boton_de_corrector = Button(botonera_concatenador, text="Corregir títulos", command=desplegar_corrector_concat)
boton_de_corrector.grid(column=1, row=3,padx=5, pady=5)
boton_de_corrector = Button(botonera_concatenador, text="Publicar por API", command=publicar)
boton_de_corrector.grid(column=0, row=5,padx=5, pady=5)
boton_de_corrector = Button(botonera_concatenador, text="Errores de publicación", command=nopubicados)
boton_de_corrector.grid(column=1, row=5,padx=5, pady=5)
bot_obtener_token = Button(botonera_concatenador, text="Obtener Token", command=revisar_tokenpresxistente)
bot_obtener_token.grid(column=0, row=6, padx=5, pady=5)


resultado = Label(frame, text="URL concatenados:")
resultado.grid(column=0, row=2, padx=5, pady=5),

frame_arbol=Frame(window)
frame_arbol.grid(column=0, row=3, padx=5, pady=5)
arb = ttk.Treeview(frame_arbol)
arb['columns']=('Título de publicación','ISBN','Imágenes','SKU','cantidad','precio','tema','título','autor','editorial')
arb.column('#0', width=0, stretch=NO)
arb.column('Título de publicación', anchor=W, width=340)
arb.column('ISBN', anchor=CENTER, width=70)
arb.column('Imágenes', anchor=CENTER, width=20)
arb.column('SKU', anchor=CENTER, width=20)
arb.column('cantidad', anchor=CENTER, width=40)
arb.column('precio', anchor=CENTER, width=40)
arb.column('tema', anchor=CENTER, width=80)
arb.column('título', anchor=CENTER, width=80)
arb.column('autor', anchor=CENTER, width=80)
arb.column('editorial', anchor=CENTER, width=80)

arb.heading('#0', text='',anchor=CENTER)
arb.heading('Título de publicación', text='Título de publicación',anchor=CENTER)
arb.heading('ISBN', text='ISBN',anchor=CENTER)
arb.heading('Imágenes', text='Imágenes',anchor=CENTER)
arb.heading('SKU', text='SKU',anchor=CENTER)
arb.heading('cantidad', text='cantidad',anchor=CENTER)
arb.heading('precio', text='precio',anchor=CENTER)
arb.heading('tema', text='tema',anchor=CENTER)
arb.heading('título', text='título',anchor=CENTER)
arb.heading('autor', text='autor',anchor=CENTER)
arb.heading('editorial', text='editorial',anchor=CENTER)
"""
cuadro_resultado = Text(frame, width=75, height=25)
cuadro_resultado.grid(column=1, row=3, padx=5, pady=5)
scroll = Scrollbar(frame, command=cuadro_resultado.yview)
scroll.grid(column=2, row=3, sticky="nsew")
cuadro_resultado.config(yscrollcommand=scroll.set)
"""


window.mainloop()