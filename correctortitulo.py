import re
import openpyxl as op
def exceliando():
	excel=op.load_workbook("nuevoenviolibrosya.xlsx")
	hoja=excel['Hoja1']
	largo=len(hoja['A'])+1
	for x in range(2,largo):
		titul=hoja.cell(row=x,column=2).value
		print(titul)
		print(type(titul))
		hoja.cell(row=x,column=2).value=corrector_titulo(titul)
		autor=hoja.cell(row=x,column=3).value
		hoja.cell(row=x,column=3).value=corrector_titulo(autor)

	excel.save("nuevoenviolibrosya.xlsx")	
def corrector_titulo(pretitulo):
	retitulo=re.search(r"(.*), ([LE][A-Z]\w?)(.*)", pretitulo)
	if retitulo == None:
		prepre=pretitulo
	else:
		titulo=(retitulo[2]+" "+retitulo[1]+" "+retitulo[3]).strip()
		prepre=titulo
	retitulo2=re.search(r"([\w 0-9]*)(/L)([\w 0-9]*)",prepre)
	if retitulo2 == None:
		return prepre.strip()
	else:
		titulo=(retitulo2[1]+retitulo2[3]).strip()
		return titulo
def corrector_autor(preautor):	
	reautor = re.search(r"([\w ]*), ([\w ]*)",preautor)
	if reautor == None:
		return (preautor.strip(),preautor.strip())
	else:	
		autor=reautor[2]+" "+reautor[1]
		apellido=reautor[1]
		return (autor, apellido)